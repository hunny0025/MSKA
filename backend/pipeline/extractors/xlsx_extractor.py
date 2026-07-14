"""
XLSX text extractor.
"""
from pipeline.extractors import ExtractionError


def extract(file_path: str) -> str:
    """Extract text from an .xlsx file row by row. Raises ExtractionError on failure."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        raise ExtractionError("openpyxl is not installed")

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        parts = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    parts.append(row_text)
        wb.close()
        result = "\n".join(parts).strip()
        if not result:
            raise ExtractionError(f"XLSX produced empty text: {file_path}")
        return result
    except ExtractionError:
        raise
    except Exception as exc:
        raise ExtractionError(f"XLSX extraction failed for {file_path}: {exc}") from exc
